# -*- coding: utf-8 -*-
import glob
import json
import re
import unicodedata
from collections import Counter

import openpyxl


NINE_DISTRICTS = [
    "上城区",
    "拱墅区",
    "西湖区",
    "滨江区",
    "钱塘区",
    "萧山区",
    "余杭区",
    "临平区",
    "富阳区",
]
REGION_PREFIX = [
    "杭州市",
    "杭州",
    "上城区",
    "拱墅区",
    "西湖区",
    "滨江区",
    "钱塘区",
    "萧山区",
    "富阳区",
    "临平区",
    "余杭区",
    "临安区",
    "桐庐县",
    "淳安县",
    "建德市",
]
CLEAN_REGIONS = ["上城", "拱墅", "西湖", "滨江", "钱塘", "萧山", "余杭", "临平", "富阳", "临安", "桐庐", "淳安", "建德"]
GEO_HEADS = ["湾", "河", "湖", "山", "江", "岛", "塘", "街道", "镇", "乡"]
SCHOOL_HEADS = ["中学", "学校", "实验", "初级", "第", "一中", "二中", "三中", "四中", "五中", "六中", "十中", "外国语", "教育集团"]
GENERIC_BLACKLIST = {
    "中学",
    "学校",
    "实验中学",
    "实验学校",
    "实验外国语学校",
    "外国语学校",
    "第一中学",
    "第二中学",
    "第三中学",
    "第四中学",
    "第五中学",
    "教育集团",
    "一中实验",
}


def clean_name(raw):
    if not raw:
        return raw
    s = str(raw).strip()
    if s.startswith("杭州市"):
        tail = s[len("杭州市"):]
        if tail in GENERIC_BLACKLIST:
            return str(raw).strip()
        s = tail
    elif s.startswith("杭州"):
        tail = s[len("杭州"):]
        if any(tail.startswith(r) for r in CLEAN_REGIONS):
            s = tail
    for r in CLEAN_REGIONS:
        for suf in ("区", "县", "市"):
            token = r + suf
            if token in s:
                candidate = s.replace(token, "")
                if candidate in GENERIC_BLACKLIST:
                    return str(raw).strip()
                s = candidate
    for r in CLEAN_REGIONS:
        if s.startswith(r):
            rest = s[len(r):]
            if not any(rest.startswith(g) for g in GEO_HEADS):
                if rest and rest not in GENERIC_BLACKLIST:
                    s = rest
            break
    return s.strip() or str(raw).strip()


def norm2(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = s.replace(" ", "").replace("\u3000", "")
    for p in REGION_PREFIX:
        if s.startswith(p):
            s = s[len(p):]
    s = s.replace("初级中学", "初中")
    return s


def variants(s):
    base = norm2(s)
    out = {base}
    if base:
        for a, b in (
            ("实验学校", "学校"),
            ("中学教育集团", "中学"),
            ("附属学校", "学校"),
            ("初级中学", "初中"),
            ("第一中学", "一中"),
            ("第二中学", "二中"),
            ("第三中学", "三中"),
            ("第四中学", "四中"),
            ("第五中学", "五中"),
            ("第六中学", "六中"),
            ("第十中学", "十中"),
            ("学校", "中学"),
            ("中学", "初中"),
        ):
            out.add(base.replace(a, b))
        extra = []
        for v in list(out):
            v_plain = re.sub(r"[（(]初中部?[)）]", "", v)
            extra.append(v_plain)
            extra.append(v_plain.replace("第一初中", "一中"))
            extra.append(v_plain.replace("第二初中", "二中"))
            extra.append(v_plain.replace("第三初中", "三中"))
            extra.append(v_plain.replace("第四初中", "四中"))
            extra.append(v_plain.replace("第五初中", "五中"))
            extra.append(v_plain.replace("镇中", "镇初中"))
            extra.append(v_plain.replace("镇", ""))
        out.update(extra)
    return out


def sim_match(n1, n2):
    if not n1 or not n2:
        return False
    if n1 == n2:
        return True
    v1, v2 = variants(n1), variants(n2)
    for a in v1:
        for b in v2:
            if a == b:
                return True
            if len(a) >= 4 and len(b) >= 4:
                if (a in b or b in a) and not (b.startswith(a) or a.startswith(b)):
                    return True
    return False


def find_key(items, district, name):
    n = norm2(name)
    for k, _ in items:
        if k[0] == district and sim_match(k[1], n):
            return k
    return None


def is_school_name(name):
    if not name:
        return False
    if not re.search(r"[\u4e00-\u9fa5]", name):
        return False
    return any(w in name for w in ("中学", "学校", "教育集团"))


path = [
    p
    for p in glob.glob(r"E:\ai\codex-project\shengxuechuzhong\*.xlsx")
    if "分配生" in p and not p.split("\\")[-1].startswith("~$")
][0]
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

# ---- 256 所名单骨架 ----
ws = wb.worksheets[14]
base_rows = []
for r in ws.iter_rows(min_col=1, max_col=6, values_only=True):
    seq = r[0]
    if seq is None or str(seq).strip() in ("", "序号"):
        continue
    district = str(r[1]).strip() if r[1] else ""
    name = str(r[2]).strip() if r[2] else ""
    if not name or not district or not is_school_name(name):
        continue
    base_rows.append(
        {
            "district": district,
            "name": name,
            "nature": str(r[3]).strip() if r[3] else "",
            "type": str(r[4]).strip() if r[4] else "",
            "note": str(r[5]).strip() if r[5] else "",
            "source": "base",
        }
    )

# ---- 分配生总表（6 区，T1-T5）----
ws = wb.worksheets[2]
sheet2_rows = []
cur_d = cur_t = None
for r in ws.iter_rows(min_col=1, max_col=11, values_only=True):
    seq = r[0]
    if seq is None or str(seq).strip() in ("", "序号"):
        continue
    d = str(r[1]).strip() if r[1] else ""
    t = str(r[2]).strip() if r[2] else ""
    name = str(r[3]).strip() if r[3] else ""
    if d:
        cur_d = d
    if t:
        m = re.search(r"\d+", t)
        cur_t = int(m.group()) if m else None
    if not name or not is_school_name(name):
        continue
    sheet2_rows.append(
        {
            "district": cur_d,
            "name": name,
            "tier": cur_t,
            "nature": str(r[4]).strip() if r[4] else "",
            "type": str(r[5]).strip() if r[5] else "",
            "cnt26": str(r[6]).strip() if r[6] is not None else "",
            "quota26": str(r[7]).strip() if r[7] is not None else "",
            "note": str(r[10]).strip() if len(r) > 10 and r[10] else "",
            "source": "sheet2",
        }
    )

# ---- 拱墅区（梯队排名 1-4）----
ws = wb.worksheets[4]
sheet4_rows = []
cur_t = None
for r in ws.iter_rows(min_col=1, max_col=10, max_row=80, values_only=True):
    seq = r[0]
    name = str(r[3]).strip() if len(r) > 3 and r[3] else ""
    if str(seq).strip() == "序号":
        continue
    if not name or not is_school_name(name):
        continue
    t = str(r[2]).strip() if len(r) > 2 and r[2] else ""
    if t:
        m = re.search(r"\d+", t)
        cur_t = int(m.group()) if m else None
    if cur_t is not None and not (1 <= cur_t <= 5):
        print("WARN 忽略异常梯队值:", cur_t, name)
        cur_t = None
    sheet4_rows.append(
        {
            "district": "拱墅区",
            "name": name,
            "tier": cur_t,
            "nature": str(r[5]).strip() if len(r) > 5 and r[5] else "",
            "type": str(r[6]).strip() if len(r) > 6 and r[6] else "",
            "cnt26": str(r[8]).strip() if len(r) > 8 and r[8] is not None else "",
            "quota26": str(r[9]).strip() if len(r) > 9 and r[9] is not None else "",
            "note": str(r[7]).strip() if len(r) > 7 and r[7] else "",
            "source": "sheet4",
        }
    )

# ---- 富阳区（名单与办学信息）----
ws = wb.worksheets[11]
sheet11_rows = []
for r in ws.iter_rows(min_col=1, max_col=6, max_row=27, values_only=True):
    seq = r[0]
    if seq is None or str(seq).strip() in ("", "序号"):
        continue
    name = str(r[2]).strip() if len(r) > 2 and r[2] else ""
    if not name or not is_school_name(name):
        continue
    sheet11_rows.append(
        {
            "district": "富阳区",
            "name": name,
            "tier": None,
            "nature": str(r[3]).strip() if len(r) > 3 and r[3] else "",
            "type": str(r[4]).strip() if len(r) > 4 and r[4] else "",
            "cnt26": "",
            "quota26": "",
            "note": str(r[5]).strip() if len(r) > 5 and r[5] else "",
            "source": "sheet11",
        }
    )

# ---- 民间榜 125 所（按重高率分档补梯队）----
ws = wb.worksheets[13]
sheet125_rows = []
for r in ws.iter_rows(min_col=1, max_col=10, values_only=True):
    district = str(r[0]).strip() if r[0] else ""
    name = str(r[1]).strip() if len(r) > 1 and r[1] else ""
    if not district or district == "区域" or not name:
        continue
    if district not in NINE_DISTRICTS:
        continue
    try:
        rate = float(r[5]) / 100.0
    except (TypeError, ValueError):
        rate = None
    tier = None
    if rate is not None:
        if rate >= 0.40:
            tier = 1
        elif rate >= 0.30:
            tier = 2
        elif rate >= 0.20:
            tier = 3
        else:
            tier = 4
    sheet125_rows.append(
        {
            "district": district,
            "name": name,
            "tier": tier,
            "nature": str(r[2]).strip() if len(r) > 2 and r[2] else "",
            "type": "",
            "cnt26": "",
            "quota26": "",
            "note": str(r[9]).strip() if len(r) > 9 and r[9] else "",
            "source": "sheet125",
        }
    )

# ---- 萧山工作表确认是否有初中名单 ----
ws = wb.worksheets[10]
cx_hits = []
for r in ws.iter_rows(min_col=1, max_col=10, max_row=236, values_only=True):
    joined = " ".join(str(c) for c in r if c is not None)
    if "初中" in joined:
        cx_hits.append([str(c) if c is not None else "" for c in r][:6])
print("萧山工作表含初中字样的行数:", len(cx_hits))
for row in cx_hits:
    print("   ", row)


# ---- 合并 ----
records_by_key = {}


def add(rec):
    k = (rec["district"], norm2(rec["name"]))
    old = records_by_key.get(k)
    if old:
        if old["tier"] is None and rec["tier"] is not None:
            old["tier"] = rec["tier"]
        for f in ("nature", "type", "cnt26", "quota26", "note"):
            if not old[f] and rec[f]:
                old[f] = rec[f]
        return
    records_by_key[k] = dict(rec)


pending125 = []


def apply_source(rows, allow_add=True):
    for item in rows:
        fk = find_key(list(records_by_key.items()), item["district"], item["name"])
        if fk:
            target = records_by_key[fk]
            if target["tier"] is None and item["tier"] is not None:
                target["tier"] = item["tier"]
            for f in ("nature", "type", "cnt26", "quota26", "note"):
                if not target[f] and item[f]:
                    target[f] = item[f]
        elif allow_add:
            add(item)
        else:
            pending125.append(item)


for b in base_rows:
    add(
        {
            "district": b["district"],
            "tier": None,
            "name": b["name"],
            "nature": b["nature"],
            "type": b["type"],
            "cnt26": "",
            "quota26": "",
            "note": b["note"],
        }
    )

apply_source(sheet2_rows)
apply_source(sheet4_rows)
apply_source(sheet11_rows)
apply_source(sheet125_rows, allow_add=False)

records = list(records_by_key.values())
for r in records:
    r.setdefault("alias", "")
SKIP_NAMES = {"采荷濮家", "启正中学"}
MANUAL_MAP = {
    "保俶塔申花实验": ("西湖区", "保俶塔申花实验", ""),
    "绿城育华亲亲": ("余杭区", "绿城育华亲亲", ""),
    "育海外国语": ("余杭区", "育海外国语", ""),
    "信达外国语": ("临平区", "信达外国语", ""),
    "英特外国语": ("余杭区", "英特外国语", ""),
    "银湖实验": ("富阳区", "银湖实验中学", "永兴分校"),
    "育才大城北": ("拱墅区", "育才大城北", ""),
    "十三中教育集团": ("西湖区", "十三中", ""),
    "钱江新城实验": ("上城区", "钱江新城实验", ""),
    "清河实验": ("上城区", "清河实验", ""),
    "翠苑中学": ("西湖区", "翠苑中学（翠苑校区、文华校区）", ""),
    "瓜沥一中": ("萧山区", "瓜沥镇第一", ""),
    "义桥实验": ("萧山区", "义桥实验", ""),
}
total_before = len(records)
records = [r for r in records if r["name"] not in SKIP_NAMES]
removed = total_before - len(records)
if removed:
    print("已移除重复校区条目:", removed)

print("=== 民间榜整合 ===")
merged_info = []
for item in pending125:
    if item["name"] in SKIP_NAMES:
        continue
    target = None
    if item["name"] in MANUAL_MAP:
        m = MANUAL_MAP[item["name"]]
        for rec in records:
            if rec["district"] == m[0] and m[1] in rec["name"] and (not m[2] or m[2] not in rec["name"]):
                target = rec
                break
    if target:
        target["alias"] = target["name"]
        target["name"] = item["name"]
        if target["tier"] is None and item["tier"] is not None:
            target["tier"] = item["tier"]
        print("  合并:", item["district"], item["name"], "->", target["alias"], "梯队", target["tier"])
        merged_info.append((item, target["alias"]))
    else:
        records.append(
            {
                "district": item["district"],
                "tier": item["tier"],
                "name": item["name"],
                "nature": item.get("nature", ""),
                "type": item.get("type", ""),
                "cnt26": "",
                "quota26": "",
                "note": item.get("note", ""),
                "alias": "",
            }
        )
        print("  新增:", item["district"], item["name"], "梯队", item["tier"])
        merged_info.append((item, ""))

print("=== 源表梯队一致性核对 ===")
issues = 0
for item in sheet2_rows + sheet4_rows:
    fk = find_key(list(records_by_key.items()), item["district"], item["name"])
    if fk:
        rec = records_by_key[fk]
        if rec["tier"] != item["tier"]:
            issues += 1
            print("  不一致:", item["district"], item["name"], "源表", item["tier"], "-> data", rec["tier"])
    else:
        issues += 1
        print("  未找到:", item["district"], item["name"])
print("一致性核对发现问题:", issues)

for r in records:
    r.pop("source", None)
    if not r.get("alias"):
        r["alias"] = r["name"]
    r["name"] = clean_name(r["name"])
name_counts = Counter(r["name"] for r in records)
for r in records:
    if name_counts[r["name"]] > 1:
        r["name"] = r["alias"]
FULL_NAME_OVERRIDES = {"十三中教育集团": "十三中教育集团（总校）杭十三中"}
for r in records:
    if r["name"] in FULL_NAME_OVERRIDES:
        r["name"] = FULL_NAME_OVERRIDES[r["name"]]
print("清理后重名回退:", sum(1 for r in records if name_counts[r["name"]] > 1))
records.sort(key=lambda x: (NINE_DISTRICTS.index(x["district"]), x["name"]))

by_d = Counter(x["district"] for x in records)
by_t = Counter(x["tier"] for x in records)
print("总校数:", len(records))
print("按区:", dict(by_d))
print("按梯队:", dict(by_t))
print("无梯队数量:", by_t.get(None, 0))
for x in records:
    if x["tier"] is None:
        print("  无梯队:", x["district"], x["name"])

# ---- 输出网页数据 ----
out_js = r"E:\ai\codex-project\shengxuechuzhong\data.js"
with open(out_js, "w", encoding="utf-8") as f:
    f.write("window.HZ_SCHOOLS = ")
    json.dump(records, f, ensure_ascii=False, indent=1)
    f.write(";\n")
print("written:", out_js)

# ---- 输出整理版 Excel ----
out_xlsx = r"E:\ai\codex-project\shengxuechuzhong\杭州9区初中梯队总表_整理版.xlsx"
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "9区初中梯队总表"
headers = ["序号", "所属区域", "梯队", "学校名称", "表格原名", "办学性质", "办学类型", "2026报考人数", "2026分配生名额", "备注"]
ws2.append(headers)
for i, rec in enumerate(records, 1):
    tier_text = "T%d" % rec["tier"] if rec["tier"] else "暂无"
    ws2.append(
        [
            i,
            rec["district"],
            tier_text,
            rec["name"],
            rec.get("alias", ""),
            rec["nature"],
            rec["type"],
            rec["cnt26"],
            rec["quota26"],
            rec["note"],
        ]
    )
widths = [6, 10, 8, 38, 38, 16, 26, 12, 14, 46]
for idx, w in enumerate(widths, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
ws_pending = wb2.create_sheet("民间榜对照")
ws_pending.append(["民间说法", "所属区域", "对应正式学校", "梯队", "备注"])
for item, full in merged_info:
    tier_txt = "T%d" % item["tier"] if item["tier"] else "暂无"
    ws_pending.append([item["name"], item["district"], full or "新增", tier_txt, item.get("note", "")])
for idx, w in enumerate([16, 10, 38, 10, 46], 1):
    ws_pending.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
wb2.save(out_xlsx)
print("written:", out_xlsx)
wb.close()
